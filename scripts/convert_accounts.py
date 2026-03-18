#!/usr/bin/env python3
"""
Конвертация Excel с лицевыми счетами в JSON для статического сайта.

Установка зависимости:
  pip install openpyxl

Пример структуры Excel (порядок колонок не важен):
  | ФИО                     | Лицевой счёт | Примечание |
  | Иванов Иван Иванович    | 1234567890   | ...        |
  | Петров Пётр Петрович    | 2468135790   | ...        |
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "data"
SOURCE_XLSX = DATA_DIR / "accounts.xlsx"
TARGET_JSON = BASE_DIR / "data" / "accounts.json"

FIO_ALIASES = {"фио", "fio"}
ACCOUNT_ALIASES = {"лицевой счет", "лицевой счёт", "счет", "счёт", "account"}


def safe_print(message: str) -> None:
    """Печать сообщения без падения в терминалах с ограниченной кодировкой."""
    encoding = sys.stdout.encoding or "utf-8"
    try:
        print(message)
    except UnicodeEncodeError:
        print(message.encode(encoding, errors="replace").decode(encoding, errors="replace"))


def normalize_header(value: Any) -> str:
    """Приводим заголовок к виду для надёжного сравнения."""
    if value is None:
        return ""

    text = str(value).strip().lower()
    text = " ".join(text.split())
    return text


def normalize_fio(value: Any) -> str:
    """Убираем лишние пробелы в ФИО."""
    if value is None:
        return ""

    return " ".join(str(value).strip().split())


def normalize_account(value: Any) -> str:
    """Сохраняем лицевой счёт как строку."""
    if value is None:
        return ""

    text = str(value).strip()

    # Excel часто хранит числа как float (например, 123.0).
    if text.endswith(".0") and text.replace(".", "", 1).isdigit():
        text = text[:-2]

    return text


def main() -> int:
    source_file = SOURCE_XLSX

    # Поддержка ситуации, когда файл назван "accounts" с другим Excel-расширением.
    if not source_file.exists():
        candidates = sorted(DATA_DIR.glob("accounts.*"))
        excel_candidates = [p for p in candidates if p.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}]

        if excel_candidates:
            source_file = excel_candidates[0]

    if not source_file.exists():
        safe_print(f"Ошибка: не найден файл {SOURCE_XLSX}")
        existing = sorted(DATA_DIR.glob("accounts.*"))
        if existing:
            safe_print("Найденные файлы с похожим именем:")
            for path in existing:
                safe_print(f"- {path.name}")
        else:
            safe_print("В папке data нет файлов accounts.*")
        return 1

    if source_file.suffix.lower() == ".xls":
        safe_print(f"Ошибка: файл {source_file.name} в старом формате .xls.")
        safe_print("Сохраните его как .xlsx и запустите скрипт снова.")
        return 1

    try:
        workbook = load_workbook(source_file, data_only=True, read_only=True)
    except Exception as error:  # noqa: BLE001
        safe_print(f"Ошибка чтения Excel: {error}")
        return 1

    sheet = workbook.active

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        safe_print("Ошибка: Excel-файл пустой.")
        return 1

    headers = [normalize_header(cell) for cell in header_row]

    fio_index = next((i for i, name in enumerate(headers) if name in FIO_ALIASES), None)
    account_index = next((i for i, name in enumerate(headers) if name in ACCOUNT_ALIASES), None)

    missing = []
    if fio_index is None:
        missing.append("ФИО (фио / ФИО / fio)")
    if account_index is None:
        missing.append("Лицевой счёт (лицевой счет / лицевой счёт / счет / счёт / account)")

    if missing:
        safe_print("Ошибка: не найдены обязательные колонки:")
        for item in missing:
            safe_print(f"- {item}")
        return 1

    accounts: list[dict[str, str]] = []

    for row in rows_iter:
        if not row:
            continue

        fio_raw = row[fio_index] if fio_index < len(row) else None
        account_raw = row[account_index] if account_index < len(row) else None

        fio = normalize_fio(fio_raw)
        account = normalize_account(account_raw)

        # Пропускаем пустые строки и строки без обязательных значений.
        if not fio or not account:
            continue

        accounts.append({"fio": fio, "account": account})

    TARGET_JSON.parent.mkdir(parents=True, exist_ok=True)
    with TARGET_JSON.open("w", encoding="utf-8") as file:
        json.dump(accounts, file, ensure_ascii=False, indent=2)

    safe_print(f"Готово: обработано {len(accounts)} записей.")
    safe_print(f"JSON обновлен: {TARGET_JSON}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
